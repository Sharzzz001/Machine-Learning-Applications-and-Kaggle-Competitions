import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# =========================
# Step 1: Read data
# =========================
df = pd.read_excel("Active_Test.xlsx", sheet_name="Cleaned_Data")

# Data cleaning
df["Remark"] = df["Remark"].fillna("").str.replace(r"Dormant since.*", "Dormant", regex=True)
df["Account"] = df["Account"].astype(str)

# Keyword mapping
keyword_map = {
    "Overdue RR": ["overdue rr"],
    "IRPQ Attestation": ["irpq attestation", "irpq"],
    "AI Proof / AI supp doc": ["ai proof", "ai supp doc", "ai documentary proof"],
    "DA": ["dia"],
    "Doc deficiency": ["doc deficiency"],
    "Physical document": ["physical document"],
    "Deceased": ["deceased", "death", "demise"],
    "Unclaimed": ["unclaimed"],
    "Bankruptcy": ["bankruptcy", "chap 11"],
    "CAC": ["cac"],
    "Initial funding": ["initial funding"],
    "Expired document": ["expired document"],
    "Restrict to 1pp": ["restrict to 1pp", "1st party payment", "first party"],
    "Block from 3Pp": ["3pp", "3rd party transfer", "3rd party txn", "third party"]
}

# =========================
# Step 2: Categorization
# =========================
def categorize_reason(row):
    reason_text = str(row["Note Block BP Status"]).lower().strip()
    matches = []

    for cat_name, kw_list in keyword_map.items():
        for kw in kw_list:
            if kw.lower() in reason_text:
                matches.append(cat_name)
                break

    if not matches:
        if reason_text == "" or reason_text == "nan":
            matches.append("Blank")
        else:
            matches.append("Uncategorized")

    return list(dict.fromkeys(matches))

df["Categories"] = df.apply(categorize_reason, axis=1)

# =========================
# Step 3: Identify exclusions
# =========================
df["is_dormant"] = df["Remark"].str.contains("dormant", case=False, na=False)
df["is_uc"] = df["Account Subtitle"].fillna("").str.contains(r"\[UC\]", case=False, na=False)
df["is_dla"] = df["Note Block BP Status"].fillna("").str.contains("dla", case=False, na=False)
df["is_nonclient"] = df["Account"].str.startswith(("3", "4", "5"))

mask_exclude = df["is_dormant"] | df["is_uc"] | df["is_dla"] | df["is_nonclient"]
mask_include = ~mask_exclude

df_included = df.loc[mask_include].copy()
df_excluded = df.loc[mask_exclude].copy()

# =========================
# Step 4: Remark combinations
# =========================
remark_combos = (
    df_included.groupby("Account")["Remark"]
    .unique()
    .reset_index()
)
remark_combos["Remark_Combo"] = remark_combos["Remark"].apply(lambda x: ", ".join(sorted(x)))

# =========================
# Step 5: Add combo info back
# =========================
df_included = df_included.merge(remark_combos[["Account", "Remark_Combo"]], on="Account", how="left")

# =========================
# Step 6: Aggregation
# =========================
# Flatten category lists for pivoting
df_included = df_included.explode("Categories")

pivot = pd.pivot_table(
    df_included,
    index="Categories",
    columns="Remark_Combo",
    values="Account",
    aggfunc=lambda x: x.nunique(),
    fill_value=0,
    margins=True,
    margins_name="Grand Total"
).reset_index()

# =========================
# Step 7: Account-level aggregation
# =========================
account_agg = (
    df_included.groupby("Account")["Categories"]
    .apply(lambda x: sorted(set(x.dropna())))
    .reset_index()
)
reason_agg = (
    df_included.groupby("Categories")["Account"]
    .nunique()
    .reset_index()
    .rename(columns={"Account": "Unique Accounts"})
)

# =========================
# Step 8: Summary values
# =========================
total_unique = df["Account"].nunique()
unique_dormant = df.loc[df["is_dormant"], "Account"].nunique()
unique_uc = df.loc[df["is_uc"], "Account"].nunique()
unique_dla = df.loc[df["is_dla"], "Account"].nunique()
unique_nonclient = df.loc[df["is_nonclient"], "Account"].nunique()
total_categorized = df_included["Account"].nunique()
total_reasons = pivot["Grand Total"].sum()

combo_counts = remark_combos["Remark_Combo"].value_counts().reset_index()
combo_counts.columns = ["Remark_Combo", "Unique Accounts"]

# =========================
# Step 9: Write to Excel
# =========================
output_file = "Categorization_Output.xlsx"
with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    pivot.to_excel(writer, sheet_name="Pivot", index=False)
    account_agg.to_excel(writer, sheet_name="Account Aggregated", index=False)
    reason_agg.to_excel(writer, sheet_name="Reason Aggregated", index=False)
    combo_counts.to_excel(writer, sheet_name="Remark Combo Distribution", index=False)

    # One sheet per remark combo
    for combo_name in remark_combos["Remark_Combo"].unique():
        combo_df = df_included[df_included["Remark_Combo"] == combo_name]
        combo_df.to_excel(writer, sheet_name=f"Combo_{remark_combos['Remark_Combo'].tolist().index(combo_name)+1}", index=False)

# =========================
# Step 10: Post-processing
# =========================
wb = load_workbook(output_file)
ws = wb["Pivot"]

# Append summary text
ws.append([])
ws.append(["Total unique accounts:", total_unique])
ws.append(["Total categorized accounts:", total_categorized])
ws.append(["Total reasons for blocked accounts:", total_reasons])
ws.append(["Unique dormant accounts:", unique_dormant])
ws.append(["Unique [UC] accounts:", unique_uc])
ws.append(["Unique DLA accounts:", unique_dla])
ws.append(["Unique Non-client/Broker accounts:", unique_nonclient])

ws.append([])
ws.append(["Distribution of unique accounts by Remark Combination:"])
for _, row in combo_counts.iterrows():
    ws.append([row["Remark_Combo"], row["Unique Accounts"]])

# Autofit columns
for wsname in wb.sheetnames:
    ws = wb[wsname]
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 2, 50)
        for cell in col:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

wb.save(output_file)
print(f"✅ Excel saved successfully as {output_file}")